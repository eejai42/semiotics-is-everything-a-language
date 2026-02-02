#!/usr/bin/env python3
"""
General-purpose Rulebook-to-XLSX+CSV transpiler.

This script reads the effortless-rulebook.json and generates:
1. An Excel workbook (rulebook.xlsx) with one worksheet per table
2. A CSV export of the LanguageCandidates tab (language_candidates.csv)
3. A CSV export of column formulas (column_formulas.csv)

The xlsx generation is identical to the xlsx substrate, but this substrate
also produces CSV files that can be used for testing without Excel dependencies.
"""

import sys
import re
import csv
from pathlib import Path

# Add project root to path for shared imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from orchestration.shared import load_rulebook, get_candidate_name_from_cwd, handle_clean_arg

# Try to import openpyxl, provide helpful error if missing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def get_table_names(rulebook):
    """Extract table names from the rulebook (excluding metadata keys)."""
    metadata_keys = {'$schema', 'model_name', 'Description', '_meta'}
    return [key for key in rulebook.keys() if key not in metadata_keys]


def build_column_map(schema):
    """Build a mapping of field names to column letters.

    Returns a dict like {'Name': 'A', 'Category': 'B', ...}
    """
    column_map = {}
    for idx, field in enumerate(schema):
        col_letter = get_column_letter(idx + 1)
        column_map[field['name']] = col_letter
    return column_map


def evaluate_formula(formula, row_data):
    """Evaluate a rulebook formula using row data.

    Handles these formula patterns:
    - String concatenation: ="Is " & {{Name}} & " a language?"
    - AND: =AND({{A}}, {{B}}, NOT({{C}}), {{D}}=2)
    - IF: =IF(condition, true_val, false_val)
    - NOT: =NOT({{Field}})
    - Equality: ={{Field}} = TRUE()

    Args:
        formula: The formula string from the rulebook
        row_data: Dict of field values for the current row

    Returns:
        The computed value
    """
    if not formula.startswith('='):
        return formula

    def get_field(name):
        """Get field value from row_data."""
        return row_data.get(name)

    def eval_expr(expr):
        """Recursively evaluate an expression."""
        expr = expr.strip()

        # Handle string concatenation (only if & is at top level)
        parts = split_by_operator(expr, ' & ')
        if len(parts) > 1:
            result = ''
            for p in parts:
                val = eval_expr(p)
                if val is not None:
                    result += str(val)
            return result if result else None

        # Handle string literals
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        # Handle TRUE/FALSE
        if expr.upper() in ('TRUE', 'TRUE()'):
            return True
        if expr.upper() in ('FALSE', 'FALSE()'):
            return False

        # Handle field references {{FieldName}}
        field_match = re.match(r'^\{\{(\w+)\}\}$', expr)
        if field_match:
            return get_field(field_match.group(1))

        # Handle numeric literals
        try:
            if '.' in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Handle AND(...)
        if expr.upper().startswith('AND('):
            inner = extract_parens(expr[3:])
            args = split_args(inner)
            for arg in args:
                val = eval_expr(arg)
                if not val:
                    return False
            return True

        # Handle NOT(...)
        if expr.upper().startswith('NOT('):
            inner = extract_parens(expr[3:])
            val = eval_expr(inner)
            return not val if val is not None else None

        # Handle IF(...)
        if expr.upper().startswith('IF('):
            inner = extract_parens(expr[2:])
            args = split_args(inner)
            if len(args) < 2:
                return None
            condition = eval_expr(args[0])
            true_val = eval_expr(args[1]) if len(args) > 1 else None
            false_val = eval_expr(args[2]) if len(args) > 2 else None
            return true_val if condition else false_val

        # Handle equality: {{Field}} = value or value = value
        if ' = ' in expr or '=' in expr:
            # Split by = but be careful of ==
            parts = re.split(r'\s*=\s*', expr, maxsplit=1)
            if len(parts) == 2:
                left = eval_expr(parts[0])
                right = eval_expr(parts[1])
                return left == right

        return None

    def extract_parens(s):
        """Extract content inside parentheses."""
        s = s.strip()
        if not s.startswith('('):
            return s
        depth = 0
        for i, c in enumerate(s):
            if c == '(':
                depth += 1
            elif c == ')':
                depth -= 1
                if depth == 0:
                    return s[1:i]
        return s[1:-1] if s.endswith(')') else s

    def split_by_operator(expr, op):
        """Split by operator, respecting parentheses and quotes."""
        parts = []
        current = ''
        depth = 0
        in_string = False
        i = 0
        while i < len(expr):
            c = expr[i]
            if c == '"':
                in_string = not in_string
                current += c
            elif not in_string:
                if c == '(':
                    depth += 1
                    current += c
                elif c == ')':
                    depth -= 1
                    current += c
                elif depth == 0 and expr[i:i+len(op)] == op:
                    parts.append(current)
                    current = ''
                    i += len(op) - 1
                else:
                    current += c
            else:
                current += c
            i += 1
        if current:
            parts.append(current)
        return parts

    def split_args(s):
        """Split comma-separated arguments, respecting parens and quotes."""
        return split_by_operator(s, ',')

    return eval_expr(formula[1:])


def get_value_for_cell(field_schema, row_data, column_map, row_num):
    """Get the value to put in a cell.

    For raw fields, returns the data value.
    For calculated fields, computes and returns the value.

    Args:
        field_schema: The field definition from the schema
        row_data: The data row dict
        column_map: Dict mapping field names to column letters
        row_num: The current row number (1-indexed)

    Returns:
        The computed value for the cell
    """
    field_name = field_schema['name']
    field_type = field_schema.get('type', 'raw')

    if field_type == 'calculated' and 'formula' in field_schema:
        # This is a calculated field - compute the value
        formula = field_schema['formula']
        return evaluate_formula(formula, row_data)
    else:
        # This is a raw field - use the data value
        value = row_data.get(field_name)

        # Handle special cases
        if value is None:
            return ''
        elif isinstance(value, bool):
            return value
        else:
            return value


def apply_header_style(cell):
    """Apply styling to header cells."""
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def apply_data_style(cell, is_calculated=False):
    """Apply styling to data cells."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

    if is_calculated:
        # Light blue background for calculated fields
        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')


def create_worksheet_from_table(workbook, table_name, table_data):
    """Create a worksheet from a rulebook table definition.

    Args:
        workbook: The openpyxl Workbook
        table_name: Name of the table (used as sheet name)
        table_data: Dict with 'schema', 'data', and optionally 'Description'

    Returns:
        The created worksheet
    """
    ws = workbook.create_sheet(title=table_name[:31])  # Excel limits sheet names to 31 chars

    schema = table_data.get('schema', [])
    data = table_data.get('data', [])

    if not schema:
        return ws

    # Build column map for formula conversion
    column_map = build_column_map(schema)

    # Create a map of field names to their type (raw vs calculated)
    field_types = {f['name']: f.get('type', 'raw') for f in schema}

    # Write header row
    for col_idx, field in enumerate(schema, 1):
        cell = ws.cell(row=1, column=col_idx, value=field['name'])
        apply_header_style(cell)

        # Set column width based on header length (minimum 12 chars)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field['name']) + 2, 12)

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):  # Start at row 2 (after header)
        for col_idx, field in enumerate(schema, 1):
            value = get_value_for_cell(field, row_data, column_map, row_idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_calculated = field.get('type') == 'calculated'
            apply_data_style(cell, is_calculated)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    return ws


def to_snake_case(name):
    """Convert PascalCase or camelCase to snake_case."""
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def export_language_candidates_csv(rulebook, output_path):
    """Export the LanguageCandidates table as a CSV file.

    This produces a flat CSV with computed values (not formulas).
    """
    # Find the LanguageCandidates table
    table_data = rulebook.get('LanguageCandidates')
    if not table_data:
        print("Warning: LanguageCandidates table not found")
        return

    schema = table_data.get('schema', [])
    data = table_data.get('data', [])

    if not schema or not data:
        print("Warning: LanguageCandidates has no schema or data")
        return

    # Build column map for formula evaluation
    column_map = build_column_map(schema)

    # Write CSV
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write header row using snake_case field names
        headers = [to_snake_case(field['name']) for field in schema]
        writer.writerow(headers)

        # Write data rows with computed values
        for row_data in data:
            row = []
            for field in schema:
                value = get_value_for_cell(field, row_data, column_map, 0)
                # Convert booleans to lowercase strings for CSV compatibility
                if isinstance(value, bool):
                    value = 'true' if value else 'false'
                elif value is None:
                    value = ''
                row.append(value)
            writer.writerow(row)

    print(f"Exported {len(data)} rows to {output_path}")


def export_column_formulas_csv(rulebook, output_path):
    """Export column formulas as a CSV file.

    This produces a CSV with columns: table_name, field_name, field_type, dag_level, formula
    """
    rows = []

    table_names = get_table_names(rulebook)

    for table_name in table_names:
        table_data = rulebook.get(table_name)
        if not isinstance(table_data, dict) or 'schema' not in table_data:
            continue

        schema = table_data.get('schema', [])

        for field in schema:
            field_name = field.get('name', '')
            field_type = field.get('type', 'raw')
            dag_level = field.get('dag_level', 0 if field_type == 'raw' else '')
            formula = field.get('formula', '')

            rows.append({
                'table_name': table_name,
                'field_name': to_snake_case(field_name),
                'field_type': field_type,
                'dag_level': dag_level,
                'formula': formula
            })

    # Write CSV
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['table_name', 'field_name', 'field_type', 'dag_level', 'formula']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Exported {len(rows)} field definitions to {output_path}")


def main():
    # Define generated files for this substrate
    GENERATED_FILES = [
        'rulebook.xlsx',
        'language_candidates.csv',
        'column_formulas.csv',
        'test-answers.json',
        'test-results.md',
    ]

    # Handle --clean argument
    if handle_clean_arg(GENERATED_FILES, "CSV substrate: Removes generated CSV files and test outputs"):
        return

    candidate_name = get_candidate_name_from_cwd()
    print(f"Generating {candidate_name} from rulebook...")

    # Load the rulebook
    try:
        rulebook = load_rulebook()
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Capture existing CSV content before regenerating
    csv_path = Path('language_candidates.csv')
    old_csv_content = None
    if csv_path.exists():
        old_csv_content = csv_path.read_text(encoding='utf-8')

    # Create workbook
    wb = Workbook()

    # Remove the default sheet created by openpyxl
    default_sheet = wb.active

    # Get all table names from the rulebook
    table_names = get_table_names(rulebook)

    if not table_names:
        print("Warning: No tables found in rulebook")
        sys.exit(1)

    print(f"Found {len(table_names)} tables: {', '.join(table_names)}")

    # Create a worksheet for each table
    for table_name in table_names:
        table_data = rulebook[table_name]

        # Skip if not a table structure (must have schema)
        if not isinstance(table_data, dict) or 'schema' not in table_data:
            print(f"  Skipping {table_name}: not a table structure")
            continue

        print(f"  Creating worksheet: {table_name}")
        schema = table_data.get('schema', [])
        data = table_data.get('data', [])

        # Count raw vs calculated fields
        raw_count = sum(1 for f in schema if f.get('type', 'raw') == 'raw')
        calc_count = sum(1 for f in schema if f.get('type') == 'calculated')

        print(f"    - {len(schema)} columns ({raw_count} raw, {calc_count} calculated)")
        print(f"    - {len(data)} data rows")

        create_worksheet_from_table(wb, table_name, table_data)

    # Remove the default empty sheet if we created other sheets
    if len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    # Save the workbook
    xlsx_path = Path('rulebook.xlsx')
    wb.save(xlsx_path)
    print(f"\nGenerated: {xlsx_path}")
    print(f"  - {len(wb.sheetnames)} worksheets")

    # Export LanguageCandidates as CSV
    export_language_candidates_csv(rulebook, csv_path)

    # Export column formulas as CSV
    formulas_path = Path('column_formulas.csv')
    export_column_formulas_csv(rulebook, formulas_path)

    # Check if CSV changed - if not, delete the xlsx
    if old_csv_content is not None:
        new_csv_content = csv_path.read_text(encoding='utf-8')
        if old_csv_content == new_csv_content:
            print(f"\nCSV unchanged - deleting {xlsx_path}")
            xlsx_path.unlink()
        else:
            print(f"\nCSV changed - keeping {xlsx_path}")

    print(f"\nDone generating {candidate_name}.")


if __name__ == "__main__":
    main()
