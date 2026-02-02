#!/usr/bin/env python3
"""
Read test answers from the xlsx substrate.

This script reads the rulebook.xlsx file and populates test-answers.json
by matching xlsx columns to JSON fields.

For cells containing Excel formulas, it evaluates the formulas using Python
since newly generated xlsx files don't have cached computed values.
"""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Add project root to path for shared imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))


def to_snake_case(name):
    """Convert PascalCase or camelCase to snake_case."""
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()


def evaluate_excel_formula_recursive(formula, cell_data, col_to_header, get_cell_value_fn):
    """Evaluate an Excel formula with recursive formula resolution.

    Args:
        formula: Excel formula string (e.g., "=$H2 = TRUE()")
        cell_data: Dict of col_letter -> raw cell value (unused, for compatibility)
        col_to_header: Dict mapping column letters to header names
        get_cell_value_fn: Callback to get cell value (evaluates formulas recursively)

    Returns:
        The computed value
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula

    def get_field_value(col_letter):
        """Get field value from column letter, evaluating formulas if needed."""
        return get_cell_value_fn(col_letter)

    def eval_expr(expr):
        """Recursively evaluate an expression."""
        expr = expr.strip()

        # Handle string concatenation
        parts = split_by_operator(expr, ' & ')
        if len(parts) > 1:
            result = ''
            for p in parts:
                val = eval_expr(p)
                # Treat None, False (from IF with no else), and empty as ""
                if val is None or val is False:
                    val = ''
                if val is not None:
                    result += str(val)
            # Return None if result is empty (no content)
            return result if result else None

        # Handle string literals
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        # Handle TRUE/FALSE
        if expr.upper() in ('TRUE', 'TRUE()'):
            return True
        if expr.upper() in ('FALSE', 'FALSE()'):
            return False

        # Handle cell references like $B2 or B2
        cell_match = re.match(r'^\$?([A-Z]+)\d+$', expr)
        if cell_match:
            return get_field_value(cell_match.group(1))

        # Handle numeric literals
        try:
            if '.' in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Handle AND(...)
        if expr.upper().startswith('AND('):
            inner = extract_parens(expr[3:])
            args = split_args(inner)
            for arg in args:
                val = eval_expr(arg)
                if not val:
                    return False
            return True

        # Handle NOT(...)
        if expr.upper().startswith('NOT('):
            inner = extract_parens(expr[3:])
            val = eval_expr(inner)
            return not val if val is not None else None

        # Handle IF(...)
        if expr.upper().startswith('IF('):
            inner = extract_parens(expr[2:])
            args = split_args(inner)
            if len(args) < 2:
                return None
            condition = eval_expr(args[0])
            true_val = eval_expr(args[1]) if len(args) > 1 else None
            # When no else clause, return empty string "" (not None or False)
            # This matches Excel behavior for string concatenation
            false_val = eval_expr(args[2]) if len(args) > 2 else ""
            return true_val if condition else false_val

        # Handle comparison operators (>, <, >=, <=) - check these before equality
        for op, fn in [(' >= ', lambda a, b: a >= b),
                       (' <= ', lambda a, b: a <= b),
                       (' > ', lambda a, b: a > b),
                       (' < ', lambda a, b: a < b),
                       ('>=', lambda a, b: a >= b),
                       ('<=', lambda a, b: a <= b),
                       ('>', lambda a, b: a > b),
                       ('<', lambda a, b: a < b)]:
            if op in expr:
                parts = expr.split(op, 1)
                if len(parts) == 2:
                    left = eval_expr(parts[0])
                    right = eval_expr(parts[1])
                    if left is not None and right is not None:
                        try:
                            return fn(left, right)
                        except TypeError:
                            return None
                    return None

        # Handle equality
        if ' = ' in expr or '=' in expr:
            parts = re.split(r'\s*=\s*', expr, maxsplit=1)
            if len(parts) == 2:
                left = eval_expr(parts[0])
                right = eval_expr(parts[1])
                return left == right

        return None

    def extract_parens(s):
        """Extract content inside parentheses."""
        s = s.strip()
        if not s.startswith('('):
            return s
        depth = 0
        for i, c in enumerate(s):
            if c == '(':
                depth += 1
            elif c == ')':
                depth -= 1
                if depth == 0:
                    return s[1:i]
        return s[1:-1] if s.endswith(')') else s

    def split_by_operator(expr, op):
        """Split by operator, respecting parentheses and quotes."""
        parts = []
        current = ''
        depth = 0
        in_string = False
        i = 0
        while i < len(expr):
            c = expr[i]
            if c == '"':
                in_string = not in_string
                current += c
            elif not in_string:
                if c == '(':
                    depth += 1
                    current += c
                elif c == ')':
                    depth -= 1
                    current += c
                elif depth == 0 and expr[i:i+len(op)] == op:
                    parts.append(current)
                    current = ''
                    i += len(op) - 1
                else:
                    current += c
            else:
                current += c
            i += 1
        if current:
            parts.append(current)
        return parts

    def split_args(s):
        """Split comma-separated arguments."""
        return split_by_operator(s, ',')

    return eval_expr(formula[1:])


def evaluate_excel_formula(formula, row_data, headers, col_to_header):
    """Evaluate an Excel formula using row data (non-recursive version).

    Converts Excel cell references like $B2 back to field values and evaluates.

    Args:
        formula: Excel formula string (e.g., "=$H2 = TRUE()")
        row_data: Dict of header -> value for the current row
        headers: List of column headers
        col_to_header: Dict mapping column letters to header names

    Returns:
        The computed value
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula

    def get_field_value(col_letter):
        """Get field value from column letter."""
        header = col_to_header.get(col_letter)
        if header and header in row_data:
            return row_data[header]
        return None

    def eval_expr(expr):
        """Recursively evaluate an expression."""
        expr = expr.strip()

        # Handle string concatenation
        parts = split_by_operator(expr, ' & ')
        if len(parts) > 1:
            result = ''
            for p in parts:
                val = eval_expr(p)
                if val is not None:
                    result += str(val)
            return result if result else None

        # Handle string literals
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        # Handle TRUE/FALSE
        if expr.upper() in ('TRUE', 'TRUE()'):
            return True
        if expr.upper() in ('FALSE', 'FALSE()'):
            return False

        # Handle cell references like $B2 or B2
        cell_match = re.match(r'^\$?([A-Z]+)\d+$', expr)
        if cell_match:
            return get_field_value(cell_match.group(1))

        # Handle numeric literals
        try:
            if '.' in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Handle AND(...)
        if expr.upper().startswith('AND('):
            inner = extract_parens(expr[3:])
            args = split_args(inner)
            for arg in args:
                val = eval_expr(arg)
                if not val:
                    return False
            return True

        # Handle NOT(...)
        if expr.upper().startswith('NOT('):
            inner = extract_parens(expr[3:])
            val = eval_expr(inner)
            return not val if val is not None else None

        # Handle IF(...)
        if expr.upper().startswith('IF('):
            inner = extract_parens(expr[2:])
            args = split_args(inner)
            if len(args) < 2:
                return None
            condition = eval_expr(args[0])
            true_val = eval_expr(args[1]) if len(args) > 1 else None
            false_val = eval_expr(args[2]) if len(args) > 2 else None
            return true_val if condition else false_val

        # Handle equality
        if ' = ' in expr or '=' in expr:
            parts = re.split(r'\s*=\s*', expr, maxsplit=1)
            if len(parts) == 2:
                left = eval_expr(parts[0])
                right = eval_expr(parts[1])
                return left == right

        return None

    def extract_parens(s):
        """Extract content inside parentheses."""
        s = s.strip()
        if not s.startswith('('):
            return s
        depth = 0
        for i, c in enumerate(s):
            if c == '(':
                depth += 1
            elif c == ')':
                depth -= 1
                if depth == 0:
                    return s[1:i]
        return s[1:-1] if s.endswith(')') else s

    def split_by_operator(expr, op):
        """Split by operator, respecting parentheses and quotes."""
        parts = []
        current = ''
        depth = 0
        in_string = False
        i = 0
        while i < len(expr):
            c = expr[i]
            if c == '"':
                in_string = not in_string
                current += c
            elif not in_string:
                if c == '(':
                    depth += 1
                    current += c
                elif c == ')':
                    depth -= 1
                    current += c
                elif depth == 0 and expr[i:i+len(op)] == op:
                    parts.append(current)
                    current = ''
                    i += len(op) - 1
                else:
                    current += c
            else:
                current += c
            i += 1
        if current:
            parts.append(current)
        return parts

    def split_args(s):
        """Split comma-separated arguments."""
        return split_by_operator(s, ',')

    return eval_expr(formula[1:])


def convert_cell_value(value):
    """Convert Excel cell value to appropriate Python type."""
    if value is None:
        return None
    elif isinstance(value, bool):
        return value
    elif isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    elif isinstance(value, str):
        if value.upper() == 'TRUE':
            return True
        elif value.upper() == 'FALSE':
            return False
        elif value.strip() == '':
            return None
        return value
    return value


def fill_null_fields_from_xlsx(xlsx_path, answers_path):
    """Read test-answers.json and fill null fields from xlsx values."""

    with open(answers_path, 'r', encoding='utf-8') as f:
        answers = json.load(f)

    if not answers:
        print("Error: test-answers.json is empty")
        sys.exit(1)

    json_fields = list(answers[0].keys())
    print(f"Found {len(json_fields)} fields in test answers")

    # Find primary key field
    pk_field = next((f for f in json_fields if f.endswith('_id')), json_fields[0])
    print(f"Primary key field: {pk_field}")

    # Load workbook
    wb = load_workbook(xlsx_path)

    # Find worksheet with primary key column
    pk_pascal = ''.join(word.capitalize() for word in pk_field.split('_'))
    ws = None
    for name in wb.sheetnames:
        sheet = wb[name]
        headers = [cell.value for cell in sheet[1] if cell.value]
        if pk_pascal in headers:
            ws = sheet
            print(f"Reading from worksheet: {name}")
            break

    if ws is None:
        print(f"Error: Could not find worksheet with column '{pk_pascal}'")
        sys.exit(1)

    # Get headers and build column mapping
    headers = [cell.value for cell in ws[1] if cell.value]
    column_map = {}  # header -> snake_case json field
    for header in headers:
        snake = to_snake_case(header)
        if snake in json_fields:
            column_map[header] = snake

    # Build column letter to header mapping for formula evaluation
    from openpyxl.utils import get_column_letter
    col_to_header = {}
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        col_to_header[col_letter] = header

    print(f"Found {len(headers)} columns, matched {len(column_map)} to JSON fields")

    # Build lookup from xlsx
    xlsx_lookup = {}
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row[:len(headers)]):
            continue

        # Build cell_data: col_letter -> raw value or formula string
        cell_data = {}
        for col_idx, cell in enumerate(row[:len(headers)]):
            col_letter = get_column_letter(col_idx + 1)
            cell_data[col_letter] = cell.value

        # Cache for evaluated formula results
        eval_cache = {}

        def get_cell_value(col_letter):
            """Get cell value, evaluating formula if needed (with caching)."""
            if col_letter in eval_cache:
                return eval_cache[col_letter]

            value = cell_data.get(col_letter)
            if isinstance(value, str) and value.startswith('='):
                # Recursively evaluate formula
                value = evaluate_excel_formula_recursive(value, cell_data, col_to_header, get_cell_value)
            else:
                value = convert_cell_value(value)

            eval_cache[col_letter] = value
            return value

        # Build the row by getting each column's value
        xlsx_row = {}
        pk_value = None
        for col_idx, header in enumerate(headers):
            col_letter = get_column_letter(col_idx + 1)
            if header in column_map:
                json_field = column_map[header]
                value = get_cell_value(col_letter)
                xlsx_row[json_field] = value
                if json_field == pk_field:
                    pk_value = value

        if pk_value is not None:
            xlsx_lookup[pk_value] = xlsx_row

    print(f"Loaded {len(xlsx_lookup)} rows from xlsx")

    # Fill null fields
    fields_filled = 0
    records_updated = 0

    for answer in answers:
        pk_value = answer.get(pk_field)
        if pk_value is None or pk_value not in xlsx_lookup:
            continue

        xlsx_row = xlsx_lookup[pk_value]
        record_updated = False

        for field in json_fields:
            if answer.get(field) is None and xlsx_row.get(field) is not None:
                answer[field] = xlsx_row[field]
                fields_filled += 1
                record_updated = True

        if record_updated:
            records_updated += 1

    print(f"Filled {fields_filled} null fields across {records_updated} records")

    with open(answers_path, 'w', encoding='utf-8') as f:
        json.dump(answers, f, indent=2)

    print(f"Updated {answers_path}")


def main():
    script_dir = Path(__file__).parent
    xlsx_path = script_dir / 'rulebook.xlsx'
    answers_path = script_dir / 'test-answers.json'

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    if not answers_path.exists():
        print(f"Error: {answers_path} not found")
        sys.exit(1)

    fill_null_fields_from_xlsx(xlsx_path, answers_path)
    print("xlsx: test-answers.json updated with values from rulebook.xlsx")


if __name__ == "__main__":
    main()
