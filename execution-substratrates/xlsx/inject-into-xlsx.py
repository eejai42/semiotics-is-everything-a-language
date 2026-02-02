#!/usr/bin/env python3
"""
General-purpose Rulebook-to-XLSX transpiler.

This script reads the effortless-rulebook.json and generates an Excel workbook
with one worksheet per table, including:
- All columns from the schema
- Data rows with raw values
- Excel formulas for calculated fields

The script is generic and works for ANY rulebook structure - it simply reads
what's defined and generates the corresponding xlsx.

Smart Update Feature:
To avoid unnecessary file changes caused by volatile functions like NOW(),
the script exports the LanguageCandidates sheet to CSV before and after
regeneration. If the content is identical, the update is rolled back.
"""

import sys
import re
import os
import csv
import shutil
from pathlib import Path
from datetime import datetime

# Add project root to path for shared imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from orchestration.shared import load_rulebook, get_candidate_name_from_cwd, handle_clean_arg

# Try to import openpyxl, provide helpful error if missing
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def get_table_names(rulebook):
    """Extract table names from the rulebook (excluding metadata keys)."""
    metadata_keys = {'$schema', 'model_name', 'Description', '_meta'}
    return [key for key in rulebook.keys() if key not in metadata_keys]


def build_column_map(schema):
    """Build a mapping of field names to column letters.

    Returns a dict like {'Name': 'A', 'Category': 'B', ...}
    """
    column_map = {}
    for idx, field in enumerate(schema):
        col_letter = get_column_letter(idx + 1)
        column_map[field['name']] = col_letter
    return column_map


def convert_formula_to_excel(formula, column_map, row_num):
    """Convert a rulebook formula to an Excel formula.

    Converts {{FieldName}} placeholders to cell references like $B2.
    The $ before the column letter makes it an absolute column reference,
    while the row number remains relative.

    Args:
        formula: The formula string from the rulebook (e.g., "={{HasSyntax}} = TRUE()")
        column_map: Dict mapping field names to column letters
        row_num: The current row number (1-indexed, accounting for header row)

    Returns:
        Excel formula string with cell references
    """
    result = formula

    # Find all {{FieldName}} patterns and replace with cell references
    pattern = r'\{\{(\w+)\}\}'

    def replace_field(match):
        field_name = match.group(1)
        if field_name in column_map:
            col_letter = column_map[field_name]
            return f'${col_letter}{row_num}'
        else:
            # Keep the placeholder if field not found (might be an error in rulebook)
            return match.group(0)

    result = re.sub(pattern, replace_field, result)
    return result


def evaluate_formula(formula, row_data):
    """Evaluate a rulebook formula using row data.

    Handles these formula patterns:
    - String concatenation: ="Is " & {{Name}} & " a language?"
    - AND: =AND({{A}}, {{B}}, NOT({{C}}), {{D}}=2)
    - IF: =IF(condition, true_val, false_val)
    - NOT: =NOT({{Field}})
    - Equality: ={{Field}} = TRUE()

    Args:
        formula: The formula string from the rulebook
        row_data: Dict of field values for the current row

    Returns:
        The computed value
    """
    if not formula.startswith('='):
        return formula

    def get_field(name):
        """Get field value from row_data."""
        return row_data.get(name)

    def eval_expr(expr):
        """Recursively evaluate an expression."""
        expr = expr.strip()

        # Handle string concatenation (only if & is at top level)
        parts = split_by_operator(expr, ' & ')
        if len(parts) > 1:
            result = ''
            for p in parts:
                val = eval_expr(p)
                if val is not None:
                    result += str(val)
            return result if result else None

        # Handle string literals
        if expr.startswith('"') and expr.endswith('"'):
            return expr[1:-1]

        # Handle TRUE/FALSE
        if expr.upper() in ('TRUE', 'TRUE()'):
            return True
        if expr.upper() in ('FALSE', 'FALSE()'):
            return False

        # Handle field references {{FieldName}}
        field_match = re.match(r'^\{\{(\w+)\}\}$', expr)
        if field_match:
            return get_field(field_match.group(1))

        # Handle numeric literals
        try:
            if '.' in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Handle AND(...)
        if expr.upper().startswith('AND('):
            inner = extract_parens(expr[3:])
            args = split_args(inner)
            for arg in args:
                val = eval_expr(arg)
                if not val:
                    return False
            return True

        # Handle NOT(...)
        if expr.upper().startswith('NOT('):
            inner = extract_parens(expr[3:])
            val = eval_expr(inner)
            return not val if val is not None else None

        # Handle IF(...)
        if expr.upper().startswith('IF('):
            inner = extract_parens(expr[2:])
            args = split_args(inner)
            if len(args) < 2:
                return None
            condition = eval_expr(args[0])
            true_val = eval_expr(args[1]) if len(args) > 1 else None
            false_val = eval_expr(args[2]) if len(args) > 2 else None
            return true_val if condition else false_val

        # Handle equality: {{Field}} = value or value = value
        if ' = ' in expr or '=' in expr:
            # Split by = but be careful of ==
            parts = re.split(r'\s*=\s*', expr, maxsplit=1)
            if len(parts) == 2:
                left = eval_expr(parts[0])
                right = eval_expr(parts[1])
                return left == right

        return None

    def extract_parens(s):
        """Extract content inside parentheses."""
        s = s.strip()
        if not s.startswith('('):
            return s
        depth = 0
        for i, c in enumerate(s):
            if c == '(':
                depth += 1
            elif c == ')':
                depth -= 1
                if depth == 0:
                    return s[1:i]
        return s[1:-1] if s.endswith(')') else s

    def split_by_operator(expr, op):
        """Split by operator, respecting parentheses and quotes."""
        parts = []
        current = ''
        depth = 0
        in_string = False
        i = 0
        while i < len(expr):
            c = expr[i]
            if c == '"':
                in_string = not in_string
                current += c
            elif not in_string:
                if c == '(':
                    depth += 1
                    current += c
                elif c == ')':
                    depth -= 1
                    current += c
                elif depth == 0 and expr[i:i+len(op)] == op:
                    parts.append(current)
                    current = ''
                    i += len(op) - 1
                else:
                    current += c
            else:
                current += c
            i += 1
        if current:
            parts.append(current)
        return parts

    def split_args(s):
        """Split comma-separated arguments, respecting parens and quotes."""
        return split_by_operator(s, ',')

    return eval_expr(formula[1:])


def get_value_for_cell(field_schema, row_data, column_map, row_num):
    """Get the value to put in a cell.

    For raw fields, returns the data value.
    For calculated fields with formulas, returns the Excel formula.

    Args:
        field_schema: The field definition from the schema
        row_data: The data row dict
        column_map: Dict mapping field names to column letters
        row_num: The current row number (1-indexed)

    Returns:
        The value or Excel formula to put in the cell
    """
    field_name = field_schema['name']
    field_type = field_schema.get('type', 'raw')

    if field_type == 'calculated' and 'formula' in field_schema:
        # This is a calculated field - use the Excel formula
        formula = field_schema['formula']
        return convert_formula_to_excel(formula, column_map, row_num)
    else:
        # This is a raw field - use the data value
        value = row_data.get(field_name)

        # Handle special cases
        if value is None:
            return ''
        elif isinstance(value, bool):
            return value
        else:
            return value


def apply_header_style(cell):
    """Apply styling to header cells."""
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def apply_data_style(cell, is_calculated=False):
    """Apply styling to data cells."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border

    if is_calculated:
        # Light blue background for calculated fields
        cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')


def create_worksheet_from_table(workbook, table_name, table_data):
    """Create a worksheet from a rulebook table definition.

    Args:
        workbook: The openpyxl Workbook
        table_name: Name of the table (used as sheet name)
        table_data: Dict with 'schema', 'data', and optionally 'Description'

    Returns:
        The created worksheet
    """
    ws = workbook.create_sheet(title=table_name[:31])  # Excel limits sheet names to 31 chars

    schema = table_data.get('schema', [])
    data = table_data.get('data', [])

    if not schema:
        return ws

    # Build column map for formula conversion
    column_map = build_column_map(schema)

    # Create a map of field names to their type (raw vs calculated)
    field_types = {f['name']: f.get('type', 'raw') for f in schema}

    # Write header row
    for col_idx, field in enumerate(schema, 1):
        cell = ws.cell(row=1, column=col_idx, value=field['name'])
        apply_header_style(cell)

        # Set column width based on header length (minimum 12 chars)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field['name']) + 2, 12)

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):  # Start at row 2 (after header)
        for col_idx, field in enumerate(schema, 1):
            value = get_value_for_cell(field, row_data, column_map, row_idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            is_calculated = field.get('type') == 'calculated'
            apply_data_style(cell, is_calculated)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    return ws


def export_sheet_to_csv(xlsx_path, sheet_name, csv_path):
    """Export a specific sheet from an xlsx file to CSV.
    
    Args:
        xlsx_path: Path to the xlsx file
        sheet_name: Name of the sheet to export
        csv_path: Path for the output CSV file
        
    Returns:
        True if export succeeded, False if sheet not found or file doesn't exist
    """
    xlsx_path = Path(xlsx_path)
    csv_path = Path(csv_path)
    
    if not xlsx_path.exists():
        print(f"  Note: {xlsx_path} does not exist (first run?)")
        return False
    
    try:
        wb = load_workbook(xlsx_path, data_only=True)  # data_only=True reads computed values
        
        # Find the sheet (case-insensitive matching)
        matching_sheet = None
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                matching_sheet = name
                break
        
        if not matching_sheet:
            print(f"  Warning: Sheet '{sheet_name}' not found in {xlsx_path}")
            wb.close()
            return False
        
        ws = wb[matching_sheet]
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                # Convert all values to strings, handling None
                writer.writerow(['' if v is None else str(v) for v in row])
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  Error exporting sheet: {e}")
        return False


def compare_csv_files(csv1_path, csv2_path):
    """Compare two CSV files content.
    
    Args:
        csv1_path: Path to first CSV file
        csv2_path: Path to second CSV file
        
    Returns:
        True if files are identical, False otherwise
    """
    csv1_path = Path(csv1_path)
    csv2_path = Path(csv2_path)
    
    if not csv1_path.exists() or not csv2_path.exists():
        return False
    
    try:
        with open(csv1_path, 'r', encoding='utf-8') as f1:
            content1 = f1.read()
        with open(csv2_path, 'r', encoding='utf-8') as f2:
            content2 = f2.read()
        return content1 == content2
    except Exception as e:
        print(f"  Error comparing files: {e}")
        return False


def cleanup_file(path):
    """Safely remove a file if it exists."""
    path = Path(path)
    if path.exists():
        path.unlink()


def compute_table_values_to_csv(rulebook, table_name, csv_path):
    """Compute values from rulebook and write to CSV.

    This is used for the "after" comparison since newly generated xlsx files
    with formulas don't have cached computed values until opened in Excel.

    Args:
        rulebook: The loaded rulebook dict
        table_name: Name of the table to export
        csv_path: Path for the output CSV file

    Returns:
        True if export succeeded, False otherwise
    """
    csv_path = Path(csv_path)

    # Find the table (case-insensitive matching)
    matching_table = None
    for name in rulebook.keys():
        if name.lower() == table_name.lower():
            matching_table = name
            break

    if not matching_table:
        print(f"  Warning: Table '{table_name}' not found in rulebook")
        return False

    table_data = rulebook[matching_table]
    if not isinstance(table_data, dict) or 'schema' not in table_data:
        print(f"  Warning: '{table_name}' is not a valid table structure")
        return False

    schema = table_data.get('schema', [])
    data = table_data.get('data', [])

    try:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Write header row
            header = [field['name'] for field in schema]
            writer.writerow(header)

            # Write data rows with computed values
            for row_data in data:
                row = []
                for field in schema:
                    field_name = field['name']
                    field_type = field.get('type', 'raw')

                    if field_type == 'calculated' and 'formula' in field:
                        # Compute the value using Python formula evaluation
                        value = evaluate_formula(field['formula'], row_data)
                    else:
                        # Raw field - use the data value
                        value = row_data.get(field_name)

                    # Convert to string for CSV
                    if value is None:
                        row.append('')
                    elif isinstance(value, bool):
                        row.append(str(value))
                    else:
                        row.append(str(value))

                writer.writerow(row)

        return True

    except Exception as e:
        print(f"  Error computing values to CSV: {e}")
        return False


def generate_workbook(rulebook, table_names):
    """Generate the workbook from the rulebook.
    
    Args:
        rulebook: The loaded rulebook dict
        table_names: List of table names to process
        
    Returns:
        The generated Workbook object
    """
    wb = Workbook()
    default_sheet = wb.active

    for table_name in table_names:
        table_data = rulebook[table_name]

        if not isinstance(table_data, dict) or 'schema' not in table_data:
            print(f"  Skipping {table_name}: not a table structure")
            continue

        print(f"  Creating worksheet: {table_name}")
        schema = table_data.get('schema', [])
        data = table_data.get('data', [])

        raw_count = sum(1 for f in schema if f.get('type', 'raw') == 'raw')
        calc_count = sum(1 for f in schema if f.get('type') == 'calculated')

        print(f"    - {len(schema)} columns ({raw_count} raw, {calc_count} calculated)")
        print(f"    - {len(data)} data rows")

        create_worksheet_from_table(wb, table_name, table_data)

    if len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    return wb


def main():
    # Define generated files for this substrate
    GENERATED_FILES = [
        'rulebook.xlsx',
        'test-answers.json',
        'test-results.md',
    ]

    # Handle --clean argument
    if handle_clean_arg(GENERATED_FILES, "XLSX substrate: Removes generated Excel workbook and test outputs"):
        return

    candidate_name = get_candidate_name_from_cwd()
    print(f"Generating {candidate_name} from rulebook...")

    # Define paths
    output_path = Path('rulebook.xlsx')
    backup_path = Path('rulebook.xlsx.backup')
    csv_before_path = Path('.language_candidates_before.csv')
    csv_after_path = Path('.language_candidates_after.csv')
    comparison_sheet = 'LanguageCandidates'  # Sheet to compare for content changes

    # Load the rulebook
    try:
        rulebook = load_rulebook()
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Get all table names from the rulebook
    table_names = get_table_names(rulebook)

    if not table_names:
        print("Warning: No tables found in rulebook")
        sys.exit(1)

    print(f"Found {len(table_names)} tables: {', '.join(table_names)}")

    # --- SMART UPDATE WORKFLOW ---
    # Step 1: Export the comparison sheet from current xlsx to CSV (baseline)
    print(f"\n--- Smart Update: Checking for actual content changes ---")
    has_existing_xlsx = output_path.exists()
    baseline_exported = False
    
    if has_existing_xlsx:
        print(f"Step 1: Exporting '{comparison_sheet}' from existing xlsx to CSV...")
        baseline_exported = export_sheet_to_csv(output_path, comparison_sheet, csv_before_path)
        
        if baseline_exported:
            print(f"  Exported baseline to: {csv_before_path}")
            
            # Step 2: Rename current xlsx to backup
            print(f"Step 2: Creating backup of existing xlsx...")
            shutil.move(str(output_path), str(backup_path))
            print(f"  Backed up to: {backup_path}")
        else:
            print(f"  Skipping smart update (could not export baseline)")
    else:
        print(f"  No existing xlsx found - this is a fresh generation")

    # Step 3: Generate the new workbook
    print(f"\nStep 3: Generating new xlsx...")
    wb = generate_workbook(rulebook, table_names)

    # Save the new workbook
    wb.save(output_path)
    print(f"\nGenerated: {output_path}")
    print(f"  - {len(wb.sheetnames)} worksheets")

    # Step 4-6: Compare and decide whether to keep or rollback
    if baseline_exported and backup_path.exists():
        print(f"\nStep 4: Computing values for '{comparison_sheet}' to CSV...")
        # Use Python formula evaluation since the new xlsx doesn't have cached computed values yet
        after_exported = compute_table_values_to_csv(rulebook, comparison_sheet, csv_after_path)

        if after_exported:
            print(f"  Computed values to: {csv_after_path}")
            
            # Step 5: Compare the CSVs
            print(f"\nStep 5: Comparing content...")
            content_changed = not compare_csv_files(csv_before_path, csv_after_path)
            
            if content_changed:
                # Content actually changed - keep the new file
                print(f"  CONTENT CHANGED - keeping new xlsx")
                print(f"\nStep 6: Cleaning up (keeping new file)...")
                cleanup_file(backup_path)
                cleanup_file(csv_before_path)
                cleanup_file(csv_after_path)
                print(f"  Removed backup and temp CSV files")
            else:
                # Content is the same - rollback to avoid unnecessary file change
                print(f"  NO CONTENT CHANGE - rolling back to preserve original file")
                print(f"\nStep 6: Rolling back...")
                cleanup_file(output_path)
                shutil.move(str(backup_path), str(output_path))
                cleanup_file(csv_before_path)
                cleanup_file(csv_after_path)
                print(f"  Restored original xlsx, removed temp files")
                print(f"\n*** XLSX NOT UPDATED (content unchanged) ***")
                return  # Exit early since we rolled back
        else:
            # Couldn't export after - just keep the new file and clean up
            print(f"  Warning: Could not export new xlsx for comparison - keeping new file")
            cleanup_file(backup_path)
            cleanup_file(csv_before_path)

    print(f"\nDone generating {candidate_name}.")


if __name__ == "__main__":
    main()
